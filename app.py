# -*- coding: utf-8 -*-
import os
import sys
import json
import io
import socket
import urllib.parse
from flask import Flask, render_template, request, jsonify, send_file
from flask_socketio import SocketIO, emit
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def get_base_dir():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

def get_writable_dir():
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        return exe_dir
    if os.environ.get('VERCEL') or not os.access(os.path.dirname(os.path.abspath(__file__)), os.W_OK):
        tmp_dir = os.path.join('/tmp', 'formation_data')
        os.makedirs(tmp_dir, exist_ok=True)
        return tmp_dir
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_dir()
WRITABLE_DIR = get_writable_dir()
DATA_DIR = os.path.join(WRITABLE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

PLAYERS_FILE = os.path.join(DATA_DIR, 'players.json')
TACTICS_FILE = os.path.join(DATA_DIR, 'tactics.json')
DEFAULT_PLAYERS_FILE = os.path.join(BASE_DIR, 'data', 'default_players.json')
DEFAULT_TACTICS_FILE = os.path.join(BASE_DIR, 'data', 'tactics.json')

def init_files():
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
    except Exception:
        pass

    if not os.path.exists(PLAYERS_FILE):
        if os.path.exists(DEFAULT_PLAYERS_FILE):
            with open(DEFAULT_PLAYERS_FILE, 'r', encoding='utf-8') as sf:
                content = sf.read()
            with open(PLAYERS_FILE, 'w', encoding='utf-8') as df:
                df.write(content)
        else:
            with open(PLAYERS_FILE, 'w', encoding='utf-8') as df:
                json.dump([], df, ensure_ascii=False)

    if not os.path.exists(TACTICS_FILE):
        if os.path.exists(DEFAULT_TACTICS_FILE):
            with open(DEFAULT_TACTICS_FILE, 'r', encoding='utf-8') as sf:
                content = sf.read()
            with open(TACTICS_FILE, 'w', encoding='utf-8') as df:
                df.write(content)

init_files()

app = Flask(__name__,
            template_folder=os.path.join(BASE_DIR, 'templates'),
            static_folder=os.path.join(BASE_DIR, 'static'))
app.config['SECRET_KEY'] = 'busan_fc_formation_secret_2026'

socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return '127.0.0.1'

def get_editor_name():
    raw = request.headers.get('X-Editor-Name') or request.args.get('editor')
    if raw:
        try:
            return urllib.parse.unquote(raw)
        except Exception:
            return raw
    return '동료 코치'

active_connections = 0

@socketio.on('connect')
def handle_connect():
    global active_connections
    active_connections += 1
    socketio.emit('users_count', {'count': active_connections})
    emit('sync_state', {
        'players': load_players(),
        'tactics': load_tactics(),
        'users_count': active_connections
    })

@socketio.on('disconnect')
def handle_disconnect():
    global active_connections
    if active_connections > 0:
        active_connections -= 1
    socketio.emit('users_count', {'count': active_connections})

@socketio.on('client_update_player')
def handle_client_update_player(data):
    editor = data.get('editor', '동료 코치')
    player = data.get('player')
    socketio.emit('player_updated', {'player': player, 'editor': editor}, include_self=False)

@socketio.on('client_update_tactics')
def handle_client_update_tactics(data):
    editor = data.get('editor', '동료 코치')
    tactics = data.get('tactics')
    socketio.emit('tactics_updated', {'tactics': tactics, 'editor': editor}, include_self=False)

def calculate_ovr(pos, stats):
    pac = stats.get('pac', 70)
    sho = stats.get('sho', 70)
    pas = stats.get('pas', 70)
    dri = stats.get('dri', 70)
    def_ = stats.get('def', 70)
    phy = stats.get('phy', 70)

    pos = (pos or 'CM').upper()
    if pos in ['ST', 'CF']:
        ovr = sho * 0.35 + pac * 0.25 + phy * 0.15 + dri * 0.15 + pas * 0.10
    elif pos in ['LW', 'RW', 'LM', 'RM']:
        ovr = pac * 0.30 + dri * 0.25 + pas * 0.20 + sho * 0.15 + phy * 0.10
    elif pos in ['CAM', 'AM']:
        ovr = pas * 0.30 + dri * 0.25 + sho * 0.20 + pac * 0.15 + phy * 0.10
    elif pos in ['CM']:
        ovr = pas * 0.25 + dri * 0.20 + phy * 0.20 + def_ * 0.15 + sho * 0.10 + pac * 0.10
    elif pos in ['CDM', 'DM']:
        ovr = def_ * 0.30 + phy * 0.25 + pas * 0.20 + dri * 0.10 + pac * 0.10 + sho * 0.05
    elif pos in ['CB']:
        ovr = def_ * 0.40 + phy * 0.30 + pac * 0.15 + pas * 0.10 + dri * 0.05
    elif pos in ['LB', 'RB', 'LWB', 'RWB']:
        ovr = pac * 0.25 + def_ * 0.25 + phy * 0.20 + pas * 0.15 + dri * 0.15
    elif pos in ['GK']:
        ovr = def_ * 0.35 + phy * 0.25 + pac * 0.15 + pas * 0.15 + dri * 0.10
    else:
        ovr = (pac + sho + pas + dri + def_ + phy) / 6.0
    return int(round(ovr))

def normalize_positions(positions_input, fallback_pos='CM'):
    if isinstance(positions_input, list):
        clean = [str(p).strip().upper() for p in positions_input if str(p).strip()]
        return clean[:3] if clean else [fallback_pos]
    elif isinstance(positions_input, str):
        parts = [p.strip().upper() for p in positions_input.replace('/', ',').split(',') if p.strip()]
        return parts[:3] if parts else [fallback_pos]
    return [fallback_pos]

def load_players():
    if not os.path.exists(PLAYERS_FILE):
        init_files()
    target_file = PLAYERS_FILE if os.path.exists(PLAYERS_FILE) else DEFAULT_PLAYERS_FILE
    try:
        with open(target_file, 'r', encoding='utf-8') as f:
            players = json.load(f)
            for p in players:
                pos = p.get('position', 'CM')
                if 'positions' not in p or not p['positions']:
                    p['positions'] = [pos]
                else:
                    p['positions'] = normalize_positions(p['positions'], pos)
                p['position'] = p['positions'][0]
                p.pop('department', None)
                s = p.get('stats', {})
                if 'total_score' not in p:
                    p['total_score'] = sum([s.get('pac', 70), s.get('sho', 70), s.get('pas', 70), s.get('dri', 70), s.get('def', 70), s.get('phy', 70)])
            return players
    except Exception:
        return []

def save_players(players):
    for p in players:
        p.pop('department', None)
    with open(PLAYERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(players, f, ensure_ascii=False, indent=2)

def load_tactics():
    if not os.path.exists(TACTICS_FILE):
        init_files()
    target_file = TACTICS_FILE if os.path.exists(TACTICS_FILE) else DEFAULT_TACTICS_FILE
    try:
        with open(target_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_tactics(tactics):
    with open(TACTICS_FILE, 'w', encoding='utf-8') as f:
        json.dump(tactics, f, ensure_ascii=False, indent=2)

def calculate_ai_lineup(formation_key):
    tactics = load_tactics()
    players = load_players()
    formConfig = tactics.get('formations', {}).get(formation_key)
    if not formConfig or not players:
        return {}

    slots = formConfig.get('slots', [])
    if len(players) < len(slots):
        return {}

    def get_slot_score(p, role):
        pos_list = p.get('positions', [p.get('position', 'CM')])
        ovr = p.get('ovr', 75)
        if role in pos_list:
            idx = pos_list.index(role)
            pos_mult = 1.10 if idx == 0 else (1.05 if idx == 1 else 1.02)
        else:
            adj_groups = [
                {'ST','CF'}, {'LW','LM','RW','RM'}, {'CAM','CM'},
                {'CDM','CM'}, {'LB','LWB'}, {'RB','RWB'}, {'CB','CDM'}, {'GK'}
            ]
            is_adj = any(role in g and any(pos in g for pos in pos_list) for g in adj_groups)
            if is_adj:
                pos_mult = 0.88
            else:
                pos_mult = 0.10 if (role == 'GK' or 'GK' in pos_list) else 0.70

        return ovr * pos_mult

    available_pids = set(p['id'] for p in players)
    assignment = {}

    # 1. Assign GK slots
    gk_slots = [s for s in slots if s['role'] == 'GK']
    for s in gk_slots:
        best_p = max(players, key=lambda p: get_slot_score(p, 'GK') if p['id'] in available_pids else -999)
        assignment[s['slotId']] = best_p['id']
        available_pids.remove(best_p['id'])

    # 2. Sort other slots by position rarity
    other_slots = [s for s in slots if s['role'] != 'GK']
    other_slots.sort(key=lambda s: sum(1 for p in players if s['role'] in p.get('positions', [])))

    for s in other_slots:
        best_p = max(players, key=lambda p: get_slot_score(p, s['role']) if p['id'] in available_pids else -999)
        assignment[s['slotId']] = best_p['id']
        available_pids.remove(best_p['id'])

    # 3. 2-opt optimization
    for _ in range(6):
        for s1 in slots:
            for s2 in slots:
                if s1['slotId'] == s2['slotId']:
                    continue
                p1 = next((p for p in players if p['id'] == assignment[s1['slotId']]), None)
                p2 = next((p for p in players if p['id'] == assignment[s2['slotId']]), None)
                if not p1 or not p2:
                    continue
                curr_score = get_slot_score(p1, s1['role']) + get_slot_score(p2, s2['role'])
                swap_score = get_slot_score(p2, s1['role']) + get_slot_score(p1, s2['role'])
                if swap_score > curr_score:
                    assignment[s1['slotId']] = p2['id']
                    assignment[s2['slotId']] = p1['id']

    return assignment

@app.route('/api/server-info', methods=['GET'])
def get_server_info():
    ip = get_local_ip()
    port = request.host.split(':')[1] if ':' in request.host else '5000'
    return jsonify({
        'local_ip': ip,
        'port': port,
        'network_url': f"http://{ip}:{port}",
        'local_url': f"http://localhost:{port}"
    })

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/players', methods=['GET'])
def get_players():
    return jsonify(load_players())

@app.route('/api/players', methods=['POST'])
def save_player():
    data = request.json
    data.pop('department', None)
    players = load_players()

    player_id = data.get('id')
    stats = data.get('stats', {'pac': 70, 'sho': 70, 'pas': 70, 'dri': 70, 'def': 70, 'phy': 70})

    positions = normalize_positions(data.get('positions', [data.get('position', 'CM')]))
    primary_pos = positions[0]

    ovr = calculate_ovr(primary_pos, stats)
    total_score = sum([stats.get('pac', 70), stats.get('sho', 70), stats.get('pas', 70), stats.get('dri', 70), stats.get('def', 70), stats.get('phy', 70)])

    data['positions'] = positions
    data['position'] = primary_pos
    data['ovr'] = ovr
    data['total_score'] = total_score

    if not player_id:
        new_id = f'p_{len(players) + 1}_{int(os.times().system*1000)}'
        data['id'] = new_id
        players.append(data)
    else:
        found = False
        for idx, p in enumerate(players):
            if p.get('id') == player_id:
                players[idx] = data
                found = True
                break
        if not found:
            players.append(data)

    save_players(players)
    editor = get_editor_name()
    socketio.emit('player_updated', {'player': data, 'editor': editor})
    return jsonify({'success': True, 'player': data})

@app.route('/api/players/<player_id>', methods=['DELETE'])
def delete_player(player_id):
    players = load_players()
    target_player = next((p for p in players if p.get('id') == player_id), None)
    p_name = target_player.get('name', '선수') if target_player else '선수'
    players = [p for p in players if p.get('id') != player_id]
    save_players(players)

    tactics = load_tactics()
    starting11 = tactics.get('starting11', {})
    for slot, pid in list(starting11.items()):
        if pid == player_id:
            starting11[slot] = None
    subs = tactics.get('substitutes', [])
    tactics['substitutes'] = [pid for pid in subs if pid != player_id]
    save_tactics(tactics)

    editor = get_editor_name()
    socketio.emit('player_deleted', {'player_id': player_id, 'player_name': p_name, 'editor': editor})
    socketio.emit('tactics_updated', {'tactics': tactics, 'editor': editor})
    return jsonify({'success': True})

@app.route('/api/tactics', methods=['GET'])
def get_tactics():
    return jsonify(load_tactics())

@app.route('/api/tactics', methods=['POST'])
def update_tactics():
    data = request.json
    save_tactics(data)
    editor = get_editor_name()
    socketio.emit('tactics_updated', {'tactics': data, 'editor': editor})
    return jsonify({'success': True})

@app.route('/api/tactics/recommend', methods=['GET'])
def get_recommendation():
    formation_key = request.args.get('formation', '4-3-3')
    recommended_lineup = calculate_ai_lineup(formation_key)
    return jsonify({
        'success': True,
        'formation': formation_key,
        'recommended11': recommended_lineup
    })

@app.route('/api/reset-default', methods=['POST'])
def reset_default():
    p_data = []
    t_data = {}
    if os.path.exists(DEFAULT_PLAYERS_FILE):
        with open(DEFAULT_PLAYERS_FILE, 'r', encoding='utf-8') as sf:
            p_data = json.load(sf)
        save_players(p_data)
    if os.path.exists(DEFAULT_TACTICS_FILE):
        with open(DEFAULT_TACTICS_FILE, 'r', encoding='utf-8') as sf:
            t_data = json.load(sf)
        save_tactics(t_data)
    editor = get_editor_name()
    socketio.emit('players_imported', {'players': p_data, 'tactics': t_data, 'editor': editor})
    return jsonify({'success': True})

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    players = load_players()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '선수단_명단_입력양식'

    ws.merge_cells('A1:Q1')
    title_cell = ws['A1']
    title_cell.value = '★ 부산시청 축구회 선수단 명단 및 능력치 입력 양식 (포지션: 드롭박스 선택 / I~N열 점수 입력시 O열 총점 및 P열 OVR 자동 계산) ★'
    title_cell.font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='0C2D48', end_color='0C2D48', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = [
        '선수ID', '등번호', '이름', '나이',
        '주포지션(1순위)', '부포지션(2순위)', '부포지션(3순위)', '주발',
        'PAC(주력)', 'SHO(슈팅)', 'PAS(패스)', 'DRI(드리블)', 'DEF(수비)', 'PHY(피지컬)',
        '총점(6개합계)', '종합평점(OVR)', '선수특징/메모'
    ]
    ws.append(headers)
    ws.row_dimensions[2].height = 24

    header_font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    total_hdr_fill = PatternFill(start_color='C65911', end_color='C65911', fill_type='solid')

    data_font = Font(name='맑은 고딕', size=10)
    bold_font = Font(name='맑은 고딕', size=10, bold=True)
    total_cell_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = total_hdr_fill if col_idx in [15, 16] else header_fill
        cell.alignment = center_align

    # Add Player Rows
    for idx, p in enumerate(players):
        row_num = idx + 3
        stats = p.get('stats', {})
        pos_list = p.get('positions', [p.get('position', 'CM')])
        p1 = pos_list[0] if len(pos_list) > 0 else 'CM'
        p2 = pos_list[1] if len(pos_list) > 1 else ''
        p3 = pos_list[2] if len(pos_list) > 2 else ''

        row = [
            p.get('id', f'p{idx+1}'),
            p.get('back_number', idx + 1),
            p.get('name', ''),
            p.get('age', 30),
            p1, p2, p3,
            p.get('foot', '오른발'),
            stats.get('pac', 70),
            stats.get('sho', 70),
            stats.get('pas', 70),
            stats.get('dri', 70),
            stats.get('def', 70),
            stats.get('phy', 70),
            f'=SUM(I{row_num}:N{row_num})',
            f'=ROUND(AVERAGE(I{row_num}:N{row_num}), 0)',
            p.get('notes', '')
        ]
        ws.append(row)
        ws.row_dimensions[row_num].height = 20

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = bold_font if col_idx in [2, 3, 15, 16] else data_font
            cell.border = thin_border
            if col_idx in [15, 16]:
                cell.fill = total_cell_fill
            cell.alignment = left_align if col_idx in [3, 17] else center_align

    # Add 25 blank rows with formulas for new players
    start_blank = len(players) + 3
    for i in range(25):
        row_num = start_blank + i
        row = [
            f'p_new_{i+1}',
            '', '', '',
            'CM', '', '', '오른발',
            '', '', '', '', '', '',
            f'=IF(COUNT(I{row_num}:N{row_num})>0, SUM(I{row_num}:N{row_num}), "")',
            f'=IF(COUNT(I{row_num}:N{row_num})>0, ROUND(AVERAGE(I{row_num}:N{row_num}), 0), "")',
            ''
        ]
        ws.append(row)
        ws.row_dimensions[row_num].height = 20
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [15, 16]:
                cell.fill = total_cell_fill
            cell.alignment = left_align if col_idx in [3, 17] else center_align

    # Dropdown validations (DataValidation)
    dv_pos = DataValidation(type='list', formula1='"ST,LW,RW,CAM,CM,CDM,LB,CB,RB,GK"', allow_blank=True)
    ws.add_data_validation(dv_pos)
    dv_pos.add(f'E3:G{start_blank + 24}')

    dv_foot = DataValidation(type='list', formula1='"오른발,왼발,양발"', allow_blank=True)
    ws.add_data_validation(dv_foot)
    dv_foot.add(f'H3:H{start_blank + 24}')

    column_widths = {
        'A': 10, 'B': 8, 'C': 12, 'D': 8,
        'E': 15, 'F': 15, 'G': 15, 'H': 10,
        'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 12,
        'O': 15, 'P': 14, 'Q': 35
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='부산시청_축구회_선수명단_입력양식.xlsx'
    )

@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '파일이 전송되지 않았습니다.'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': '선택된 파일이 없습니다.'}), 400

    try:
        df = pd.read_excel(file, header=1)
        if '이름' not in df.columns:
            file.seek(0)
            df = pd.read_excel(file, header=0)

        new_players = []
        for idx, row in df.iterrows():
            name = str(row.get('이름', '')).strip()
            if not name or name == 'nan' or '★' in name:
                continue

            pos1 = str(row.get('주포지션(1순위)', row.get('주포지션', 'CM'))).strip().upper()
            pos2 = str(row.get('부포지션(2순위)', '')).strip().upper()
            pos3 = str(row.get('부포지션(3순위)', '')).strip().upper()

            pos_candidates = []
            for p_item in [pos1, pos2, pos3]:
                if p_item and p_item != 'NAN':
                    for sub in p_item.replace('/', ',').split(','):
                        sub_clean = sub.strip()
                        if sub_clean and sub_clean not in pos_candidates:
                            pos_candidates.append(sub_clean)

            positions = pos_candidates[:3] if pos_candidates else ['CM']
            primary_pos = positions[0]

            pac = int(row.get('PAC(주력)', 70)) if pd.notna(row.get('PAC(주력)')) else 70
            sho = int(row.get('SHO(슈팅)', 70)) if pd.notna(row.get('SHO(슈팅)')) else 70
            pas = int(row.get('PAS(패스)', 70)) if pd.notna(row.get('PAS(패스)')) else 70
            dri = int(row.get('DRI(드리블)', 70)) if pd.notna(row.get('DRI(드리블)')) else 70
            def_ = int(row.get('DEF(수비)', 70)) if pd.notna(row.get('DEF(수비)')) else 70
            phy = int(row.get('PHY(피지컬)', 70)) if pd.notna(row.get('PHY(피지컬)')) else 70

            stats = {'pac': pac, 'sho': sho, 'pas': pas, 'dri': dri, 'def': def_, 'phy': phy}
            total_score = pac + sho + pas + dri + def_ + phy
            ovr = calculate_ovr(primary_pos, stats)

            pid = str(row.get('선수ID', '')).strip()
            if not pid or pid == 'nan':
                pid = f'p_imp_{idx+1}_{int(os.times().system*1000)}'

            p = {
                'id': pid,
                'back_number': int(row.get('등번호', idx + 1)) if pd.notna(row.get('등번호')) else idx + 1,
                'name': name,
                'age': int(row.get('나이', 30)) if pd.notna(row.get('나이')) else 30,
                'positions': positions,
                'position': primary_pos,
                'secondary_positions': positions[1:],
                'foot': str(row.get('주발', '오른발')).strip() if pd.notna(row.get('주발')) else '오른발',
                'stats': stats,
                'total_score': total_score,
                'ovr': ovr,
                'notes': str(row.get('선수특징/메모', '')).strip() if pd.notna(row.get('선수특징/메모')) else ''
            }
            new_players.append(p)

        if not new_players:
            return jsonify({'success': False, 'error': '엑셀에서 유효한 선수 데이터를 찾을 수 없습니다.'}), 400

        save_players(new_players)
        editor = get_editor_name()
        socketio.emit('players_imported', {'players': new_players, 'count': len(new_players), 'editor': editor})
        return jsonify({'success': True, 'count': len(new_players)})
    except Exception as e:
        return jsonify({'success': False, 'error': f'엑셀 처리 중 오류 발생: {str(e)}'}), 500

if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, allow_unsafe_werkzeug=True)
